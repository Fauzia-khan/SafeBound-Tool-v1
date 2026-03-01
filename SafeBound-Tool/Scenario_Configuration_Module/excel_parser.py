from openpyxl import load_workbook


def parse_scenario_tags(excel_filename, scenario_row):
    """
    Extract scenario tags and light condition from an Excel file.

    Returns:
        tag_data (dict) – categories and their selected tags
        light_condition (str or None) – "Day" or "Night"
    """

    scenario_row += 1  # Skip header (your original logic)

    wb = load_workbook(excel_filename)
    ws = wb.active

    tag_data = {}
    light_condition = None

    # Column ranges
    actor_column_start, actor_column_end = 5, 10
    weather_column_start, weather_column_end = 11, 14
    light_column_start, light_column_end = 15, 17
    behaviour_column_start, behaviour_column_end = 18, 32
    road_topology_column_start, road_topology_column_end = 33, 35
    scenario_group_column = 38

    # Actors
    for i in range(actor_column_start, actor_column_end):
        if ws.cell(row=scenario_row, column=i).value == 1:
            tag_data.setdefault("Actors", []).append(ws.cell(row=1, column=i).value.split('_')[1])

    # Weather
    for i in range(weather_column_start, weather_column_end):
        if ws.cell(row=scenario_row, column=i).value == 1:
            tag_data.setdefault("Weather", []).append(ws.cell(row=1, column=i).value.split('_')[1])

    # Light
    for i in range(light_column_start, light_column_end):
        if ws.cell(row=scenario_row, column=i).value == 1:
            value = ws.cell(row=1, column=i).value.split('_')[1]
            tag_data.setdefault("Light", []).append(value)
            light_condition = value

    # Driving Maneuver
    for i in range(behaviour_column_start, behaviour_column_end):
        if ws.cell(row=scenario_row, column=i).value == 1:
            tag_data.setdefault("Driving Maneuver", []).append(ws.cell(row=1, column=i).value.split('_')[1])

    # Road Topology
    for i in range(road_topology_column_start, road_topology_column_end):
        if ws.cell(row=scenario_row, column=i).value == 1:
            tag_data.setdefault("Road Topology", []).append(ws.cell(row=1, column=i).value.split('_')[1])

    # Scenario Group
    scenario_group = ws.cell(row=scenario_row, column=scenario_group_column).value
    if scenario_group:
        tag_data["Scenario Group"] = [scenario_group]

    return tag_data, light_condition
def get_weather_type(tag_data):
    if "Weather" not in tag_data:
        return "dry"

    weather_tags = [tag.lower() for tag in tag_data["Weather"]]

    for w in ["dry", "rain", "snow", "fog"]:
        if w in weather_tags:
            return w

    return "dry"
