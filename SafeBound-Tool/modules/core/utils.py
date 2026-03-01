import os
import pandas as pd
from openpyxl import load_workbook
from Scenario_Selection_Module import CATALOG_SCENARIOS_PATH

def delete_scenario_from_excel_file(selected_indexes):

    excel_file = CATALOG_SCENARIOS_PATH

    wb = load_workbook(excel_file)
    ws = wb.active

    for i in range(3, ws.max_row + 1):
        if ws.cell(row=i, column=2).value in selected_indexes:
            ws.delete_rows(i, 2)

        wb.save(excel_file)


def save_to_excel(name, description, actor, weather, light, maneuver, road_topology, scenario_id, image_file_path,catalog):
    # Define multi-level columns (Category -> Subcategory)
    columns = [
        ('Scenario ID', ''),
        ('Name', ''),
        ('Description', ''),
        ('Actors', 'Vehicle'),
        ('Actors', 'Pedestrian'),
        ('Actors', 'Motorcyclist'),
        ('Actors', 'Pedal cyclist'),
        ('Actors', 'Animal'),
        ('Actors', 'Other'),
        ('Weather', 'Dry'),
        ('Weather', 'Rain'),
        ('Weather', 'Snow'),
        ('Weather', 'Fog'),
        ('Light', 'Day'),
        ('Light', 'Dark'),
        ('Light', 'Twilight'),
        ('Driving Maneuver', 'Driving Straight'),
        ('Driving Maneuver', 'Negotiating a Curve'),
        ('Driving Maneuver', 'Turning Left'),
        ('Driving Maneuver', 'Passing or Overtaking Another Vehicle'),
        ('Driving Maneuver', 'Merging/Changing Lanes'),
        ('Driving Maneuver', 'Stopped in Roadway'),
        ('Driving Maneuver', 'Other Maneuver'),
        ('Driving Maneuver', 'Turning Right'),
        ('Driving Maneuver', 'Decelerating in Road'),
        ('Driving Maneuver', 'Starting in Road'),
        ('Driving Maneuver', 'Making a U-turn'),
        ('Driving Maneuver', 'Backing Up'),
        ('Driving Maneuver', 'Parked'),
        ('Driving Maneuver', 'Leaving a Parking Position'),
        ('Driving Maneuver', 'Entering a Parking Position'),
        ('Road Topology', 'With Signalized Junction'),
        ('Road Topology', 'Non-Signalized Junction'),
        ('Road Topology', 'Non-Junction'),
        ('Image Path', ''),
        ('Catalog', ''),
    ]

    # Scenario data: 1 if the category matches, 0 otherwise
    scenario_data = [
        [
            scenario_id,  # Placeholder for Scenario ID
            name,
            description,
            1 if actor == 'Vehicle' else 0,  # Actors
            1 if actor == 'Pedestrian' else 0,
            1 if actor == 'Motorcyclist' else 0,
            1 if actor == 'Cyclist' else 0,
            1 if actor == 'Animal' else 0,
            1 if actor == 'Other' else 0,
            1 if weather == 'Dry' else 0,  # Weather
            1 if weather == 'Rain' else 0,
            1 if weather == 'Snow' else 0,
            1 if weather == 'Fog' else 0,
            1 if light == 'Day' else 0,  # Light
            1 if light == 'Dark' else 0,
            1 if light == 'Twilight' else 0,
            1 if maneuver == 'Driving Straight' else 0, # Driving Maneuver
            1 if maneuver == 'Negotiating a Curve' else 0,
            1 if maneuver == 'Turning Left' else 0,
            1 if maneuver == 'Passing or Overtaking Another Vehicle' else 0,
            1 if maneuver == 'Merging/Changing Lanes' else 0,
            1 if maneuver == 'Stopped in Roadway' else 0,
            1 if maneuver == 'Other Maneuver' else 0,
            1 if maneuver == 'Turning Right' else 0,
            1 if maneuver == 'Decelerating in Road' else 0,
            1 if maneuver == 'Starting in Road' else 0,
            1 if maneuver == 'Making a U-turn' else 0,
            1 if maneuver == 'Backing Up' else 0,
            1 if maneuver == 'Parked' else 0,
            1 if maneuver == 'Leaving a Parking Position' else 0,
            1 if maneuver == 'Entering a Parking Position' else 0,
            1 if road_topology == 'With Signalized Junction' else 0,  # Road Topology
            1 if road_topology == 'Non-Signalized Junction' else 0,
            1 if road_topology == 'Non-Junction' else 0,
            image_file_path,
            catalog,
        ]
    ]

    # Create MultiIndex columns
    multi_index_columns = pd.MultiIndex.from_tuples(columns, names=['Category', 'Subcategory'])

    # Convert the scenario data to a pandas DataFrame
    df = pd.DataFrame(scenario_data, columns=multi_index_columns)

    # Define the file path for the Excel file
    excel_file = CATALOG_SCENARIOS_PATH
    os.makedirs(os.path.dirname(excel_file), exist_ok=True)
    # ✅ Safe write
    # ✅ If file exists, append to it
    if os.path.exists(excel_file):
        workbook = load_workbook(excel_file)

        # ✅ Safely get or create 'Scenarios' sheet
        if 'Scenarios' in workbook.sheetnames:
            sheet = workbook['Scenarios']
        else:
            sheet = workbook.create_sheet('Scenarios')

        start_row = sheet.max_row if sheet.max_row > 1 else 1
        print("start_row", start_row)

        # Append data using pandas writer (modern pandas — no writer.book assignment!)
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, sheet_name='Scenarios', index=True, header=False, startrow=start_row)

    # ✅ Otherwise, create a new file and write headers
    else:
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, sheet_name='Scenarios', index=True, header=True)

    print('✅ Success: Scenario saved to Excel successfully!')
    remap_excel_indexes(excel_file)

def remap_excel_indexes(excel_file, start_from=None):
    df = pd.read_excel(excel_file)
    wb = load_workbook(excel_file)
    ws = wb.active

    if start_from == None:

        for i in range(3, df.shape[0]+1):
            ws[f"A{i+1}"] = i-2

    else:
        for i in range(start_from, df.shape[0]+1):
            ws[f"A{i + 1}"] = i - 2

    wb.save(excel_file)