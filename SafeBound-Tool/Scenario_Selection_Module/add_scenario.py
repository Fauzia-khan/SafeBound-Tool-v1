
from PyQt5.QtWidgets import QWidget, QPushButton, QComboBox, QLabel, QLineEdit, QTextEdit, QGridLayout, QFileDialog, QMessageBox, QDialog
from PyQt5.QtCore import pyqtSignal,Qt
from PyQt5.QtCore import QTimer

from Scenario_Selection_Module import CATALOG_SCENARIOS_PATH
from openpyxl import load_workbook
from modules.core.utils import save_to_excel

# Add Scenario Window
class AddScenarioWindow(QDialog):
    scenario_added = pyqtSignal()

    def __init__(self, catalog, parent=None):

        super().__init__(parent)
        self.catalog = catalog
        self.setWindowTitle(f"Add Scenario to {self.catalog}")
        self.setGeometry(400, 200, 400, 400)
        self.setWindowModality(Qt.ApplicationModal)
        self.setStyleSheet("""
                    QWidget {
                        background-color: #e9f0fa;          /* Light blue background */
                        font-family: 'Segoe UI', Arial, sans-serif;
                        font-size: 10.5pt;
                        color: #222;
                    }
                    QLabel {
                        font-weight: normal;
                    }
                    QLineEdit, QTextEdit, QComboBox {
                        background-color: #ffffff;
                        border: 1px solid #8c8c8c;
                        border-radius: 3px;
                        padding: 3px;
                    }
                    QPushButton {
                        background-color: qlineargradient(
                            x1:0, y1:0, x2:0, y2:1,
                            stop:0 #f9f9f9, stop:1 #e0e0e0
                        );
                        border: 1px solid #8c8c8c;
                        border-radius: 4px;
                        padding: 5px 10px;
                        font-weight: normal;
                    }
                    QPushButton:hover {
                        background-color: #f0f0f0;
                    }
                    QPushButton:pressed {
                        background-color: #dcdcdc;
                    }
                """)

        layout = QGridLayout()
        self.setLayout(layout)

        # Labels and Inputs
        self.name_label = QLabel('Scenario Name:')
        self.name_input = QLineEdit()
        layout.addWidget(self.name_label, 0, 0)
        layout.addWidget(self.name_input, 0, 1)

        self.desc_label = QLabel('Description:')
        self.desc_input = QTextEdit()
        layout.addWidget(self.desc_label, 1, 0)
        layout.addWidget(self.desc_input, 1, 1)
        ####
        self.catalog_label = QLabel('Catalog Name:')
        self.catalog_input = QLineEdit()
        layout.addWidget(self.catalog_label, 2, 0)
        layout.addWidget(self.catalog_input, 2, 1)

        self.actor_label = QLabel('Actors:')
        self.actor_combo = QComboBox()
        self.actor_combo.addItems(['Animal','Motorcyclist', 'Pedalcyclist','Pedestrian','Vehicle' ])
        layout.addWidget(self.actor_label, 3, 0)
        layout.addWidget(self.actor_combo, 3, 1)

        self.weather_label = QLabel('Weather:')
        self.weather_combo = QComboBox()
        self.weather_combo.addItems(['Dry','Fog', 'Snow', 'Rain'])
        layout.addWidget(self.weather_label, 4, 0)
        layout.addWidget(self.weather_combo, 4, 1)

        self.light_label = QLabel('Light Condition:')
        self.light_combo = QComboBox()
        self.light_combo.addItems(['Day', 'Dark','Twilight'])
        layout.addWidget(self.light_label, 5, 0)
        layout.addWidget(self.light_combo, 5, 1)

        self.maneuver_label = QLabel('Driving Maneuver:')
        self.maneuver_combo = QComboBox()
        self.maneuver_combo.addItems([
            'Decelerating',
            'Driving Straight',
            'Entering a Parking Position',
            'Leaving a Parking Position',
            'Merging/Changing Lanes',
            'Negotiating a Curve',
            'Overtaking Another Vehicle',
            'Parked',
            'Reversing',
            'Starting in Road',
            'Stopped in Roadway',
            'Turning Left',
            'Turning Right',
            'U-turn',])
        layout.addWidget(self.maneuver_label, 6, 0)
        layout.addWidget(self.maneuver_combo, 6, 1)

        self.topology_label = QLabel('Road Topology:')
        self.topology_combo = QComboBox()
        self.topology_combo.addItems(['With Signalized Junction', 'Non-Signalized Junction', 'Non-Junction'])
        layout.addWidget(self.topology_label, 7, 0)
        layout.addWidget(self.topology_combo, 7, 1)

        # Image upload button
        self.image_label = QLabel('Scenario Image:')
        self.image_button = QPushButton('Upload Image')
        self.image_button.clicked.connect(self.upload_image)
        self.image_path = None  # Holds the path to the selected image
        layout.addWidget(self.image_label, 8, 0)
        layout.addWidget(self.image_button, 8, 1)

        # Add Scenario Button
        self.add_button = QPushButton("Add Scenario")
        self.add_button.clicked.connect(self.add_scenario)
        layout.addWidget(self.add_button, 9, 0, 1, 2)

    def upload_image(self):
        """Open file dialog to select an image"""
        self.image_path, _ = QFileDialog.getOpenFileName(self, 'Select Image', '', 'Image Files (*.png *.jpg *.jpeg)')
        if self.image_path:
            QMessageBox.information(self, 'Image Selected', f'Image path: {self.image_path}')

    def add_scenario(self):
        name = self.name_input.text()
        description = self.desc_input.toPlainText()
        catalog = self.catalog_input.text()
        actor = self.actor_combo.currentText()
        weather = self.weather_combo.currentText()
        light = self.light_combo.currentText()
        maneuver = self.maneuver_combo.currentText()
        road_topology = self.topology_combo.currentText()

        print("checkpoint1, catalog:", catalog, "self.catalog:", self.catalog)

        if not name or not description:
            QMessageBox.warning(self, "Input Error", "Please provide a name and description.")
            return

        # Map catalog to unique code
        catalog_code_map = {
            "US": "SC1",
            "Singapore": "SC2",
            "Europe": "SC3",
            "Other": "SC4"
        }

        if len(self.catalog_input.text()) != 0:
            print("no data has entered.")
            catalog_code = 'SC' + str(catalog)
        else:

            if self.catalog.lower() in ['us', 'singapore', 'europe', 'other']:
                print("abc")
                catalog_code = catalog_code_map.get(self.catalog)
            else:
                print("2..")
                catalog_code = 'SC' + str(self.catalog)

        print("catalog code:", catalog_code)

        # Generate a unique scenario ID based on catalog
        # Generate the scenario ID


        wb = load_workbook(CATALOG_SCENARIOS_PATH)

        ws = wb.active

        _count = 0

        CATALOG_FOUND_FLAG = False


        for i in range(3, ws.max_row + 1):
            _temp_scenario_id = ws.cell(row=i, column=2).value

            if _temp_scenario_id == None:
                continue

            current_index_catalog_code = _temp_scenario_id.split('-')[0]

            if catalog_code != current_index_catalog_code:
                # another catalog's data
                continue

            elif catalog_code == current_index_catalog_code:
                _scenario_index = _temp_scenario_id.split('-')[1]
                _scenario_index = int(_scenario_index.split('S')[-1])

                if _scenario_index > _count:
                    _count = _scenario_index

                CATALOG_FOUND_FLAG = True
            else:
                quit("Raising error.")

        if CATALOG_FOUND_FLAG == False:
            _count = 0 # no scenario found. assigning a new index.

        _count = _count + 1
        scenario_id = f"{catalog_code}-S{_count}"

        image_file_path = self.image_path

        save_to_excel(name, description, actor, weather, light, maneuver, road_topology, scenario_id, image_file_path,catalog)

        QMessageBox.information(self, "Success", "Scenario added successfully!")

        self.accept()  # ✅ close the dialog first

        # ✅ Emit signal *after* the dialog has closed
        QTimer.singleShot(100, lambda: self.scenario_added.emit())

