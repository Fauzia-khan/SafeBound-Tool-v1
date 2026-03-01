from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QPushButton, QLabel, QFormLayout, QMessageBox, QDialog, QScrollArea, QCheckBox
from openpyxl import load_workbook
from modules.constants import scenarios_excel_file_name, catalog_code_map
from modules.core.utils import delete_scenario_from_excel_file

class ViewScenariosWindow(QDialog):
    def __init__(self, catalog):
        super().__init__()
        self.catalog = catalog
        self.setWindowTitle(f"Scenarios in {self.catalog} Catalog")
        self.setGeometry(300, 200, 600, 400)
        self.setStyleSheet("background-color: #e9f0fa;")
        layout = QVBoxLayout()

        # Scrollable Area
        scroll_area = QScrollArea(self)
        scroll_area.setWidgetResizable(True)
        content_widget = QWidget()
        scroll_layout = QVBoxLayout(content_widget)

        self.checkboxes = []  # Store checkbox references for deletion

        wb = load_workbook(scenarios_excel_file_name)
        ws = wb.active

        scenarios = []

        for i in range(3, ws.max_row + 1):
            _temp_scenario_id = ws.cell(row=i, column=2).value

            if _temp_scenario_id == None:
                continue

            _catalog_id = _temp_scenario_id.split('-')[0]

            if catalog_code_map[self.catalog] != _catalog_id:
                continue

            each_scenario = [ws.cell(row=i, column=2).value, #scenario_id
                             ws.cell(row=i, column=3).value, #scenario_name
                             ws.cell(row=i, column=4).value, #scenario_description
                             None,#image_path
                             ]

            scenarios.append(each_scenario)


        if len(scenarios) > 0: # scenario is available in current catalog.
            for scenario in scenarios:
                scenario_id, name, description, image_path = scenario[0], scenario[1], scenario[2], scenario[3]

                # Form layout for each scenario entry
                form_layout = QFormLayout()

                # Checkbox for selection
                checkbox = QCheckBox()
                self.checkboxes.append((checkbox, scenario_id))  # Store checkbox and associated scenario_id
                form_layout.addRow(checkbox)

                # Scenario Details
                form_layout.addRow("ID:", QLabel(scenario_id))
                form_layout.addRow("Name:", QLabel(name))
                form_layout.addRow("Description:", QLabel(description))

                # Display Image if exists
                if image_path:
                    image_item = QLabel()
                    pixmap = QPixmap(image_path).scaled(100, 100)
                    image_item.setPixmap(pixmap)
                    form_layout.addRow("Image:", image_item)
                else:
                    form_layout.addRow("Image:", QLabel("No Image"))

                scroll_layout.addLayout(form_layout)

            scroll_area.setWidget(content_widget)
            layout.addWidget(scroll_area)

            # Delete Selected Button
            self.delete_button = QPushButton("Delete Selected Scenarios")
            self.delete_button.clicked.connect(self.delete_selected_scenarios)
            layout.addWidget(self.delete_button)

        else:
            layout.addWidget(QLabel("No scenarios available in this catalog."))

        self.setLayout(layout)

    def delete_selected_scenarios(self):
        """Delete selected scenarios from the database."""
        selected_ids = [scenario_id for checkbox, scenario_id in self.checkboxes if checkbox.isChecked()]

        if not selected_ids:
            QMessageBox.warning(self, "No Selection", "Please select at least one scenario to delete.")
            return

        try:
            print("Following scenarios are going to be deleted:\n", selected_ids)
            delete_scenario_from_excel_file(selected_ids)

            QMessageBox.information(self, "Success", "Selected scenarios deleted successfully.")
            self.close()  # Close the dialog after deletion
        except Exception as e:
            QMessageBox.warning(self, "Error", f"An error occurred: {e}")
