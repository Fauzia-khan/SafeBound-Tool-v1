
from openpyxl.reader.excel import load_workbook
from Scenario_Selection_Module import USER_SELECTED_SCENARIOS_PATH, CATALOG_SCENARIOS_PATH, FORMULATED_SCENARIO_GROUPS_PATH, DUPLICATE_SCENARIO_REMOVAL_PATH,USER_SELECTED_ODD_PATH,SELECTED_SCENARIO_BASEDon_ODD_PATH
from modules.core.utils import remap_excel_indexes
import os

class SelectScenariosBasedOnOdd:
    def __init__(self):

        self.create_excel_file()
        self.get_scenarios_using_ODD_file()

    def create_excel_file(self):
        source_file = DUPLICATE_SCENARIO_REMOVAL_PATH
        destination_file = SELECTED_SCENARIO_BASEDon_ODD_PATH
        #wb = load_workbook(os.getcwd()+'/'+source_file)
        wb = load_workbook(DUPLICATE_SCENARIO_REMOVAL_PATH)
        wb.save(destination_file)

    def get_scenarios_using_ODD_file(self):

        print("you are in function.")
        #### Step1: GET ODD FILE AND ACCESS DATA

        # ODD FILE
        wb_ODD = load_workbook(USER_SELECTED_ODD_PATH)
        ws_ODD = wb_ODD.active

        ODD_file_main_information_row = 8 #7th row has the data.

        # TARGET FILE
        wb_target = load_workbook(SELECTED_SCENARIO_BASEDon_ODD_PATH)
        ws_target = wb_target.active


        """
        FIRST FILTER: WEATHER CONDITION 
        """
        weather_snow, weather_rain, weather_fog, weather_dry = False, False, False, False
        active_weathers = []
        for i in range(16, 20):
            weather_info = ws_ODD.cell(row=ODD_file_main_information_row, column=i).value
            if weather_info == 1:
                if i == 16: #snow
                    weather_snow = True
                    active_weathers.append('snow')
                elif i == 17:
                    weather_rain = True
                    active_weathers.append('rain')
                elif i == 18:
                    weather_fog = True
                    active_weathers.append('foggy')
                elif i == 19:
                    weather_dry = True
                    active_weathers.append('dry')

                #weather_row = ODD_file_main_information_row - 4
                #weather = ws_ODD.cell(row=weather_row, column=i).value

        print("Snow: ", weather_snow, "Rain:", weather_rain, "Fog:", weather_fog, "Dry:", weather_dry)

        target_file_weather_column_indices = {'dry': 11, 'rain': 12, 'snow': 13, 'foggy': 14}

        if active_weathers:

            weather_column_indices_to_filter = [target_file_weather_column_indices[weather.lower()] for weather in active_weathers]

            # FILTER OUT FROM THE SCENARIOS
            for i in range(ws_target.max_row, 2, -1): # 3. Come back to the line.
                keep_row = False
                for column_index in weather_column_indices_to_filter:
                    cell_value = ws_target.cell(row=i, column=column_index).value
                    if cell_value == 1:
                        keep_row = True
                        break # If at least one column is True, keep the row.

                if not keep_row: # If none is 1, delete the row.
                    ws_target.delete_rows(i)

        else:
            print("No weather data entered. No rows will be filtering out based on the weather.")



        """
        2nd FILTER: LIGHT CONDITION 
        """

        light_Light, light_Dark = False, False

        active_lights = []

        for i in range(20, 22):
            light_info = ws_ODD.cell(row=ODD_file_main_information_row, column=i).value
            if light_info == 1:
                if i == 20:
                    light_Light = True
                    active_lights.append('light')
                elif i == 21:
                    light_Dark = True
                    active_lights.append('dark')

                #light_row = ODD_file_main_information_row - 4
                #light = ws_ODD.cell(row=light_row, column=i).value

        print("Light:", light_Light, "Dark:", light_Dark)


        # FILTERING: LIGHT

        target_file_light_column_indices = {'light': 15, 'dark': 16, 'twilight': 17}

        if active_lights:

            light_column_indices_to_filter = [target_file_light_column_indices[light.lower()] for light in active_lights]

            for i in range(ws_target.max_row, 2, -1): #3.come back to the line.
                keep_row = False
                for column_index in light_column_indices_to_filter:

                    cell_value = ws_target.cell(row=i, column=column_index).value

                    if cell_value == 1: #delete the row if there is a 0 in the column.
                        keep_row = True
                        break

                if not keep_row:
                    ws_target.delete_rows(i)

        else:
            print("No light data entered. No rows will be filtering out based on the Light.")



        """
        3rd FILTER: ROAD TOPOLOGY 
        """


        road_roundabouts, road_nonjunction, road_signalized, road_non_signalized = False, False, False,False
        active_roads = []
        for i in range(31, 35):
            road_topology_info = ws_ODD.cell(row=ODD_file_main_information_row, column=i).value
            if road_topology_info == 1:
                if i == 31:
                    road_roundabouts = True
                    active_roads.append('roundabouts')
                if i == 32:
                    road_nonjunction = True
                    active_roads.append('non-junction')
                elif i == 33:
                    road_signalized = True
                    active_roads.append('signalized')
                elif i == 34:
                    road_non_signalized = True
                    active_roads.append('non_signalized')

                #road_topology_row = ODD_file_main_information_row - 3 # for signalized and non signalized intersections
                #road_topology = ws_ODD.cell(row=road_topology_row, column=i).value

        print("Road Roundabout:", road_roundabouts, "Road Non-junction:", road_nonjunction, "Road Signalized:", road_signalized, "Road Non Signalized:", road_non_signalized)

        # FILTERING: ROAD TOPOLOGY

        target_file_road_column_indices = {'signalized': 33, 'non_signalized':34, "non-junction":  35}

        if active_roads:

            road_column_indices_to_filter = [target_file_road_column_indices[road.lower()] for road in active_roads]

            for i in range(ws_target.max_row, 2, -1):
                keep_row = False
                for column_index in road_column_indices_to_filter:

                    cell_value = ws_target.cell(row=i, column=column_index).value

                    if cell_value == 1:
                        keep_row = True
                        break

                if not keep_row: #If none is 1, delete the row.
                    ws_target.delete_rows(i)
        else:
            print("No road data entered. No rows will be filtering out based on the road topology.")



        """
        4th FILTER: ACTOR 
        """

        actor_animal, actor_pedestrian, actor_cyclist, actor_vehicle = False, False, False, False

        active_actors = []
        for i in range(8, 13):
            actor_info = ws_ODD.cell(row=ODD_file_main_information_row, column=i).value
            if actor_info == 1:

                if i == 8:
                    actor_animal = True
                    active_actors.append('animal')

                elif i == 9:
                    actor_pedestrian = True
                    active_actors.append('pedestrian')

                elif i == 10:
                    actor_cyclist = True
                    active_actors.append('cyclist')

                elif i == 11:
                    actor_cyclist = True
                    active_actors.append('cyclist')

                elif i == 12:
                    actor_vehicle = True
                    active_actors.append('vehicle')

                #animal_row = ODD_file_main_information_row - 3 # animal
                #human_row = ODD_file_main_information_row - 2 # pedestrian, cyclist, motor
                #vehicle_row = ODD_file_main_information_row - 3 # vehicle

        print("Actors:  Animal", actor_animal, " Pedestrian:", actor_pedestrian, " Cyclist:", actor_cyclist)

        target_file_actor_column_indices = {'vehicle': 5, 'pedestrian':6, 'motorcyclist':7, 'cyclist':8, 'animal':9}

        if active_actors:

            actor_column_indices_to_filter = [target_file_actor_column_indices[actor.lower()] for actor in active_actors]

            for i in range(ws_target.max_row, 2, -1):
                keep_row = False
                for column_index in actor_column_indices_to_filter:
                    cell_value = ws_target.cell(row=i, column=column_index).value
                    if cell_value == 1:
                        keep_row = True
                        break

                if not keep_row:
                    ws_target.delete_rows(i)


        else:
            print("No actor data entered. No rows will be filtering out based on the actors.")



        """
        SAVE EXCEL FILE
        """

        ## !! SAVE FILE !! ##
        wb_target.save(SELECTED_SCENARIO_BASEDon_ODD_PATH)
        ## !! ADJUST INDEXES !! ##
        remap_excel_indexes(SELECTED_SCENARIO_BASEDon_ODD_PATH, start_from=2)
        print("scenarios are selected based on selected ODD, see file ODD_selected_scenarios.")