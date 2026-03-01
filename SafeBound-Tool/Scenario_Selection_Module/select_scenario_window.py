import os
import pandas as pd
from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QPushButton, QLabel, QFormLayout,
    QMessageBox, QDialog, QScrollArea, QCheckBox
)
from modules.core.utils import delete_scenario_from_excel_file
from openpyxl import load_workbook
from Scenario_Configuration_Module.scenario_parameter_configuration_window import ViewInformationWindow
from Scenario_Selection_Module import (
    SELECTED_SCENARIOS_US_PATH,
    SELECTED_SCENARIOS_EU_PATH
)
from functools import partial


class SelectScenarioWindow(QDialog):
    def __init__(self, dataset_name):
        super().__init__()
        self.dataset_name = dataset_name  # ✅ FIXED
        self.setWindowTitle("List of Selected Scenarios")
        self.setGeometry(300, 200, 500, 300)
        self.setStyleSheet("background-color: #e9f0fa;")
        self.setMinimumSize(500, 500)

        # Scroll area setup
        scroll_area = QScrollArea(self)
        scroll_area.setWidgetResizable(True)
        content_widget = QWidget()
        scroll_layout = QVBoxLayout(content_widget)

        self.checkboxes = []

        # Select correct scenario file
        if self.dataset_name.lower() == "eu":
            excel_filename = SELECTED_SCENARIOS_EU_PATH
        else:
            excel_filename = SELECTED_SCENARIOS_US_PATH

        self.excel_filename = excel_filename  # ✅ VERY IMPORTANT FIX

        # Load data
        df = pd.read_excel(excel_filename)

        wb = load_workbook(excel_filename)
        ws = wb.active

        scenarios = []

        for i in range(2, ws.max_row + 1):
            each_scenario = [
                ws.cell(row=i, column=2).value,  # ID
                ws.cell(row=i, column=3).value,  # Name
                ws.cell(row=i, column=4).value,  # Description
                ws.cell(row=i, column=36).value, # Image
                ws.cell(row=i, column=38).value, # Scenario Group
                ws.cell(row=i, column=44).value  # Priority
            ]
            scenarios.append(each_scenario)

        # Build UI list
        scenario_view_counter = 0
        for scenario in scenarios:
            scenario_view_counter += 1

            scenario_id, name, description, image_path, scenario_group, priority = scenario

            form_layout = QFormLayout()

            # checkbox
            checkbox = QCheckBox()
            self.checkboxes.append((checkbox, scenario_id))
            form_layout.addRow(checkbox)

            form_layout.addRow("ID:", QLabel(str(scenario_id)))
            form_layout.addRow("Name:", QLabel(str(name)))
            form_layout.addRow("Description:", QLabel(str(description)))
            form_layout.addRow("ScenarioGroup:", QLabel(str(scenario_group)))
            form_layout.addRow("Priority:", QLabel(str(priority)))

            # Image
            if image_path:
                img_label = QLabel()
                pixmap = QPixmap(os.getcwd() + '/images/' + image_path).scaled(100, 100)
                img_label.setPixmap(pixmap)
                form_layout.addRow("Image:", img_label)
            else:
                form_layout.addRow("Image:", QLabel("No Image"))

            # View button
            view_btn = QPushButton("View")
            view_btn.setFixedWidth(80)
            view_btn.clicked.connect(partial(self.view_button_callback, scenario_view_counter))
            form_layout.addRow("", view_btn)

            scroll_layout.addLayout(form_layout)

        scroll_area.setWidget(content_widget)
        main_layout = QVBoxLayout()
        main_layout.addWidget(scroll_area)
        self.setLayout(main_layout)

    # -------------------
    # VIEW BUTTON ACTION
    # -------------------
    def view_button_callback(self, scenario_row):

        print("View for Scenario:", scenario_row)

        self.view_information_window = ViewInformationWindow(
            excel_filename=self.excel_filename,
            scenario_row=scenario_row
        )
        self.view_information_window.exec_()
