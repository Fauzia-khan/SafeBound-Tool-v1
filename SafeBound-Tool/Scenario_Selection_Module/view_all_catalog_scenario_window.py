import os
import sys
from tabnanny import check
from PyQt5.QtGui import QPixmap, QStandardItemModel
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QPushButton, QLabel, QFormLayout, QMessageBox, QDialog, QScrollArea, QCheckBox,QComboBox,QHBoxLayout
from modules.core.utils import delete_scenario_from_excel_file
from modules.constants import scenarios_excel_file_name
from openpyxl import load_workbook, Workbook
import shutil
from PyQt5.QtCore import Qt,  pyqtSignal
from Scenario_Selection_Module import USER_SELECTED_SCENARIOS_PATH, CATALOG_SCENARIOS_PATH
from PyQt5.QtCore import pyqtSignal, Qt
from PyQt5.QtGui import QStandardItemModel
from PyQt5.QtWidgets import QComboBox
import traceback;
class CheckableComboBox(QComboBox):
    selection_changed = pyqtSignal()  # Seçim değiştiğinde sinyal yayar

    def __init__(self):
        super(CheckableComboBox, self).__init__()
        self.view().pressed.connect(self.handle_item_pressed)
        self.setModel(QStandardItemModel(self))
        self.setEditable(True)
        self.lineEdit().setReadOnly(True)  # Kullanıcı manuel olarak değiştiremesin

    def handle_item_pressed(self, index):
        """Seçilen öğenin işaret durumunu değiştir."""
        item = self.model().itemFromIndex(index)

        # İşaretleme durumunu değiştir
        if item.checkState() == Qt.Checked:
            item.setCheckState(Qt.Unchecked)
        else:
            item.setCheckState(Qt.Checked)

        #self.update_display_text()  # Seçimlere göre metni güncelle
        self.selection_changed.emit()  # Seçim değişti sinyalini yay

    def get_checked_items(self):
        """Seçili öğelerin listesini döndür."""
        checked_items = []
        for i in range(self.count()):
            item = self.model().item(i)
            if item.checkState() == Qt.Checked:
                checked_items.append(item.text())
        return checked_items

    def update_display_text(self):

        checked_items = self.get_checked_items()
        display_text = ", ".join(checked_items) if checked_items else "Choose a Catalog"
        self.lineEdit().setText(display_text)  # Görünen metni güncelle


# View All Scenarios Window (similar to ViewScenariosWindow but for all catalogs)
class ViewAllScenariosWindow(QDialog):
    def __init__(self):
        super().__init__()
        """
        First button: List of All Scenarios
        """
        self.selected_scenarios = set()
        self.setWindowTitle("Choose Catalog/Scenarios")
        self.setGeometry(300, 200, 600, 400)
        self.setStyleSheet("background-color: #e9f0fa;")
        layout = QVBoxLayout()
        self.catalogs = set()

        self.checkboxes = []
        '''
        up_layout = QHBoxLayout()
        
        # Catalog Combo (top bar)
        self.catalog_combo_top = QComboBox(self)
        self.ADD_ITEMS_TO_CATALOG_COMBO()
        self.catalog_combo_top.setFixedSize(120, 35)
        self.catalog_combo_top.setCurrentText("Other")
        self.catalog_combo_top.currentIndexChanged.connect(self.update_flag)

        # Flag label
        self.flag_label = QLabel(self)
        self.flag_label.setFixedSize(50, 50)
        self.flag_label.setScaledContents(True)
        self.update_flag(self.catalog_combo_top.currentIndex())

        # Add Scenario button
        self.add_button = QPushButton("Add Scenario", self)
        self.add_button.setFixedSize(140, 35)
        self.add_button.clicked.connect(self.add_scenario)
        self.add_button.clicked.connect(self.add_scenario)

        # Add them to the top layout
        up_layout.addWidget(QLabel("Catalog:"))
        up_layout.addWidget(self.catalog_combo_top)
        up_layout.addWidget(self.flag_label)
        up_layout.addStretch()
        up_layout.addWidget(self.add_button)

        # Add the top layout to the main layout
        layout.addLayout(up_layout)
        '''
        # === Top Section: Add Scenario + Catalog combo (centered together) ===
        top_container = QVBoxLayout()
        top_container.setAlignment(Qt.AlignHCenter | Qt.AlignTop)
        top_container.setContentsMargins(0, 10, 0, 10)  # control top spacing

        # --- Horizontal layout for Add Scenario + Catalog combo ---
        top_row = QHBoxLayout()
        top_row.setSpacing(20)  # space between Add button and Catalog combo

        # Add Scenario button
        self.add_button = QPushButton("Add Scenario", self)
        self.add_button.setFixedSize(140, 30)
        self.add_button.clicked.connect(self.add_scenario)
        top_row.addWidget(self.add_button)

        # Catalog area (label + combo + flag)
        catalog_layout = QHBoxLayout()
        catalog_layout.setSpacing(8)

        catalog_label = QLabel("Catalog:")
        catalog_layout.addWidget(catalog_label)

        self.catalog_combo_top = QComboBox(self)
        self.ADD_ITEMS_TO_CATALOG_COMBO()
        self.catalog_combo_top.setFixedSize(140, 30)
        self.catalog_combo_top.setCurrentText("Other")
        self.catalog_combo_top.currentIndexChanged.connect(self.update_flag)
        catalog_layout.addWidget(self.catalog_combo_top)

        self.flag_label = QLabel(self)
        self.flag_label.setFixedSize(45, 45)
        self.flag_label.setScaledContents(True)
        self.update_flag(self.catalog_combo_top.currentIndex())
        catalog_layout.addWidget(self.flag_label)

        # Combine both into the same row
        top_row.addLayout(catalog_layout)

        # Add the horizontal row to top container
        top_container.addLayout(top_row)

        # Add some space below before "Choose a Catalog" section
        top_container.addSpacing(3)  # ⬅️ adjust this for perfect spacing below

        # Add this top section to the main layout
        layout.addLayout(top_container)

        # Scrollable Area
        scroll_area = QScrollArea(self)
        scroll_area.setWidgetResizable(True)
        content_widget = QWidget()
        scroll_area.setWidget(content_widget)
        self.scroll_layout = QVBoxLayout(content_widget)

        #self.checkboxes = []  # Store checkbox references for deletion

        # Get Scenario Data
        self.scenarios = self.load_scenarios()




        self.catalogs = self.get_unique_catalogs(self.scenarios)

        self.catalog_label = QLabel("Choose a Catalog:")
        self.catalog_combo = CheckableComboBox() #QComboBox()
        for catalog in self.catalogs:
            self.catalog_combo.addItem(catalog)
            item = self.catalog_combo.model().item(self.catalog_combo.count()-1)
            item.setCheckState(Qt.Unchecked)
        self.catalog_combo.selection_changed.connect(self.update_scenarios)

        #print("SELF CATALOGS:", self.catalogs, type(self.catalogs))

        #list_converted_catalogs = list(self.catalogs)

        """self.catalog_combo.addItem("All")  # Option to show all scenarios
        item = self.catalog_combo.model().item(i+1, 0)
        item.setCheckState(Qt.Unchecked)
        """
        #self.catalog_combo.addItems(self.catalogs)
        self.catalog_combo.setEditable(False)
        self.catalog_combo.setDuplicatesEnabled(False)
        self.catalog_combo.setInsertPolicy(QComboBox.NoInsert)


        layout.addWidget(self.catalog_label)
        layout.addWidget(self.catalog_combo)
        layout.addWidget(scroll_area)



        self.update_scenarios()

        # Buttons
        self.delete_button = QPushButton("Delete Selected Scenarios")
        self.delete_button.clicked.connect(self.delete_selected_scenarios)

        self.save_button = QPushButton("Save Scenarios by Selected Catalog")

        # IF USER PICKS ONE SCENARIO APPLY FILTER ACCORDING TO SCENARIO or APPLY DIRECTLY CATALOG BASED IF CLICKED
        if len(self.get_selected_scenarios()) == 0:
            print("No scenario is selected. Saving based on the catalog!")
            self.save_button.clicked.connect(self.save_by_catalog)
        else:
            print("Saving based on the selected scenarios.")

            self.save_button.clicked.connect(self.save_by_scenario)

        layout.addWidget(self.delete_button)
        layout.addWidget(self.save_button)

        self.setLayout(layout)


    def get_unique_catalogs(self, scenarios):
        #print(scenarios)

        return set(s['catalog'] for s in scenarios if s['catalog'])

    def load_scenarios(self):

        import os

        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(BASE_DIR, "catalog_scenarios.xlsx")

        wb = load_workbook(file_path)
        #wb = load_workbook("catalog_scenarios.xlsx")
        ws = wb.active
        scenarios = []
        for i in range(3, ws.max_row + 1):
            scenario = {
                "id": ws.cell(row=i, column=2).value,
                "name": ws.cell(row=i, column=3).value,
                "description": ws.cell(row=i, column=4).value,
                "catalog": ws.cell(row=i, column=37).value,
                "image_path": ws.cell(row=i, column=36).value
            }
            scenarios.append(scenario)

        return scenarios

    def get_selected_scenarios(self):
        selected_scenarios = []
        for checkbox, scenario_id in self.checkboxes:
            if checkbox.isChecked():
                selected_scenarios.append(scenario_id)
        return selected_scenarios

    def update_scenarios(self):
        selected_catalogs = self.catalog_combo.get_checked_items()

        # Scroll layout içeriğini tamamen temizle
        while self.scroll_layout.count():
            child = self.scroll_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
            elif child.layout():
                while child.layout().count():
                    inner_child = child.layout().takeAt(0)
                    if inner_child.widget():
                        inner_child.widget().deleteLater()
                child.layout().deleteLater()

        self.checkboxes = []  # CheckBox listesini sıfırla

        # Seçili kataloglara göre filtrelenen senaryoları ekrana yükle
        for scenario in self.scenarios:
            if not selected_catalogs or scenario["catalog"] in selected_catalogs:
                form_layout = QFormLayout()

                checkbox = QCheckBox()
                if scenario["id"] in self.selected_scenarios:
                    checkbox.setChecked(True)  # Daha önce seçilmişse işaretli göster

                checkbox.stateChanged.connect(
                    lambda state, scenario_id=scenario["id"]: self.update_selected_scenarios(state, scenario_id)
                )
                self.checkboxes.append((checkbox, scenario["id"]))
                form_layout.addRow(checkbox)

                form_layout.addRow("ID:", QLabel(scenario["id"]))
                form_layout.addRow("Name:", QLabel(scenario["name"]))
                form_layout.addRow("Description:", QLabel(scenario["description"]))
                form_layout.addRow("Catalog:", QLabel(scenario["catalog"]))
                #print('CHECKPINT 11')
                if scenario["image_path"]:
                    image_item = QLabel()
                    print('image path: ', os.getcwd() + 'images/' + scenario["image_path"])
                    pixmap = QPixmap(os.getcwd() + '/images/' + scenario["image_path"]).scaled(100, 100, Qt.KeepAspectRatio)
                    image_item.setPixmap(pixmap)
                    form_layout.addRow("Image:", image_item)
                else:
                    form_layout.addRow("Image:", QLabel("No Image"))

                self.scroll_layout.addLayout(form_layout)

    def update_selected_scenarios(self, state, scenario_id):
        if state == Qt.Checked:
            self.selected_scenarios.add(scenario_id)
        else:
            self.selected_scenarios.discard(scenario_id)

    def delete_selected_scenarios(self):
        """Delete selected scenarios from the database."""
        selected_ids = [scenario_id for checkbox, scenario_id in self.checkboxes if checkbox.isChecked()]

        if not selected_ids:
            QMessageBox.warning(self, "No Selection", "Please select at least one scenario to delete.")
            return

        delete_scenario_from_excel_file(selected_indexes=selected_ids)

        QMessageBox.information(self, "Success", "Selected scenarios deleted successfully.")
        self.close()  # Close the dialog after deletion

    def save_by_scenario(self):
        if not self.selected_scenarios:
            QMessageBox.warning(self, "No Selection", "Please select at least one scenario to save.")
            return



        new_wb = load_workbook(CATALOG_SCENARIOS_PATH)
        new_ws = new_wb.active


        for i in range(new_ws.max_row, 2, -1):
            scenario_id = new_ws.cell(row=i, column=2).value
            if scenario_id not in self.selected_scenarios:
                new_ws.delete_rows(i)


        new_wb.save(USER_SELECTED_SCENARIOS_PATH)

        QMessageBox.information(self, "Success", f"Scenarios saved successfully to user_selected_scenarios_from_catalog.xlsx.")

    def ADD_ITEMS_TO_CATALOG_COMBO(self):
        """Populate the top combo with catalog items."""
        items = ["Other", "US", "Singapore", "Other"]
        self.catalog_combo_top.addItems(items)

    def update_flag(self, index):
        """Update the flag emoji when the catalog changes."""
        current_text = self.catalog_combo_top.itemText(index)

        flag_emojis = {
            "US": "🇺🇸",
            "Singapore": "🇸🇬",

            "Other": "🌐"
        }

        # Set emoji as label text
        self.flag_label.setText(flag_emojis.get(current_text, "🌐"))
        self.flag_label.setAlignment(Qt.AlignCenter)
        self.flag_label.setStyleSheet("font-size: 28px;")  # make emoji nice and big

    def add_scenario(self):
        try:
            from Scenario_Selection_Module.add_scenario import AddScenarioWindow

            selected_catalog = self.catalog_combo_top.currentText()

            dialog = AddScenarioWindow(selected_catalog, self)
            dialog.scenario_added.connect(self._on_scenario_added)
            dialog.exec_()  # Modal – blocks until closed

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open Add Scenario window:\n{str(e)}")

    def _on_scenario_added(self):
        """Called when a new scenario is added successfully."""
        print("DEBUG: _on_scenario_added() triggered!")
        self.scenarios = self.load_scenarios()
        self.update_scenarios()

    def save_by_catalog(self):
        """Save scenarios related to the selected catalog in a new Excel file."""
        #selected_catalog = self.catalog_combo.currentText()
        selected_catalogs = self.catalog_combo.get_checked_items()

        print("selected_catalog:", self.catalog_combo.currentText())
        print("info:", self.catalog_combo.get_checked_items())

        if len(selected_catalogs) == 0:
            QMessageBox.warning(self, "No Catalog Selected", "Please select at least one catalog.")
            return

        import os
        from openpyxl import load_workbook

        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(BASE_DIR, "catalog_scenarios.xlsx")

        wb = load_workbook(file_path)
        #wb = load_workbook('catalog_scenarios.xlsx')
        ws = wb.active

        # Remove all rows that don't belong to the selected catalog
        rows_to_keep = [ws[1], ws[2]]  # Headers and sub-headers (modify as needed)
        for i in range(3, ws.max_row + 1):
            if ws.cell(row=i, column=37).value in selected_catalogs:
                rows_to_keep.append(ws[i])

        ws.delete_rows(3, ws.max_row - 2)  # Delete all rows except headers
        for row in rows_to_keep[2:]:
            ws.append([cell.value for cell in row])

        wb.save(USER_SELECTED_SCENARIOS_PATH)
        QMessageBox.information(self, "Success", f"Scenarios related to catalog '{selected_catalogs}' have been saved.")
        try:
            from Scenario_Selection_Module.formulate_scenario_groups import formulate_scenario_groups
            formulate_scenario_groups()  # This will process filtered_scenarios.xlsx
            #QMessageBox.information(self, "Grouping Done", "Scenario groups have been formulated successfully.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error while grouping scenarios:\n{str(e)}")
        try:
            from Scenario_Selection_Module.scenarios_duplicate_removal import remove_duplicate_scenarios
            from Scenario_Selection_Module import FORMULATED_SCENARIO_GROUPS_PATH
            remove_duplicate_scenarios(FORMULATED_SCENARIO_GROUPS_PATH)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error while removing duplicates:\n{str(e)}")

            """if scenarios:
                     for scenario in scenarios:
                         scenario_id, name, description, catalog, image_path = scenario[0], scenario[1], scenario[2], scenario[3], scenario[4]

                         form_layout = QFormLayout()

                         checkbox = QCheckBox()
                         self.checkboxes.append((checkbox, scenario_id))  # Store checkbox and associated scenario_id
                         form_layout.addRow(checkbox)

                         form_layout.addRow("ID:", QLabel(scenario_id))
                         form_layout.addRow("Name:", QLabel(name))
                         form_layout.addRow("Description:", QLabel(description))
                         form_layout.addRow("Catalog:", QLabel(catalog))

                         if image_path:
                             #print(image_path)
                             image_item = QLabel()
                             pixmap = QPixmap(image_path).scaled(100, 100)
                             image_item.setPixmap(pixmap)
                             form_layout.addRow("Image:", image_item)
                         else:
                             form_layout.addRow("Image:", QLabel("No Image"))

                         scroll_layout.addLayout(form_layout)


                     layout.addWidget(scroll_area)

                     # Delete Selected Button
                     self.delete_button = QPushButton("Delete Selected Scenarios")
                     self.delete_button.clicked.connect(self.delete_selected_scenarios)
                     layout.addWidget(self.delete_button)

                     # Select Catalog and Save
                     self.save_button = QPushButton("Save Scenarios by Selected Catalog")
                     self.save_button.clicked.connect(self.save_by_catalog)
                     layout.addWidget(self.save_button)

                 else:
                     layout.addWidget(QLabel("No scenarios available."))
                 """